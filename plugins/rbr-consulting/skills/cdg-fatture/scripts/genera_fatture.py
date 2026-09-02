import os
import sys
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- risoluzione portabile dei file condivisi ---
BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.environ.get('CDG_HOME') or BASE)
import master_sheet as ms

OUTPUT_FILE = "FATTURE.xlsx"   # generato nella cartella corrente

if len(sys.argv) < 2:
    print("Uso: python genera_fatture.py <file_csv>")
    sys.exit(1)

csv_file = sys.argv[1]

# --- Leggi CSV ---
csv = pd.read_csv(csv_file, sep=';', encoding='utf-8')

def clean(val):
    return str(val).replace("'", '').strip() if isinstance(val, str) else val

csv = csv.map(clean)

def parse_amount(val):
    if isinstance(val, str):
        val = re.sub(r'\s', '', val).lstrip('0') or '0'
        val = val.replace(',', '.')
        try:
            return float(val)
        except:
            return 0.0
    return val

csv['Imponibile'] = csv['Imponibile/Importo (totale in euro)'].apply(parse_amount)
csv['Iva']        = csv['Imposta (totale in euro)'].apply(parse_amount)
csv.loc[csv['Tipo documento'].str.lower().str.contains('nota di credito', na=False), 'Imponibile'] *= -1

# --- Carica master categorie (Google Sheet condiviso) ---
cat_map = ms.load_master_map()

def get_cat(fornitore):
    return cat_map.get(ms.norm(fornitore))

csv['Categoria'] = csv['Denominazione fornitore'].apply(get_cat)

# --- Fornitori non mappati: aggiungi al master come "Altro" ---
nuovi = csv[csv['Categoria'].isna()]['Denominazione fornitore'].unique()
if len(nuovi) > 0:
    aggiunti = ms.append_new_suppliers(nuovi, categoria='Altro')
    print(f"Fornitori non mappati ({len(nuovi)}), aggiunti al master condiviso come 'Altro' ({len(aggiunti)} nuovi su Sheet):")
    for f in sorted(nuovi):
        print(f"  {f}")
    csv['Categoria'] = csv['Categoria'].fillna('Altro')
else:
    print("Tutti i fornitori già mappati.")

# --- Costruisci risultato ---
result = pd.DataFrame({
    'Data emissione':          csv['Data emissione'],
    'DenominazioneFornitore':  csv['Denominazione fornitore'],
    'NumeroDocumento':         csv['Numero fattura / Documento'],
    'TipoDocumento':           csv['Tipo documento'],
    'Imponibile':              csv['Imponibile'],
    'Iva':                     csv['Iva'],
    'Lordo':                   csv['Imponibile'] + csv['Iva'],
    'Categoria':               csv['Categoria'],
})

result['_sort'] = pd.to_datetime(result['Data emissione'], dayfirst=True)
result = result.sort_values('_sort', ascending=False).drop(columns='_sort').reset_index(drop=True)

# --- Scrivi FATTURE.xlsx ---
wb = Workbook()
ws = wb.active
ws.title = 'Fatture'
headers = list(result.columns)
thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    cell.fill = PatternFill('solid', start_color='1F4E79')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
ws.row_dimensions[1].height = 18

for row_idx, row in result.iterrows():
    for col_idx, (col, val) in enumerate(row.items(), 1):
        cell = ws.cell(row=row_idx+2, column=col_idx, value=val)
        cell.font = Font(name='Arial', size=10)
        cell.border = border
        if col in ('Imponibile', 'Iva', 'Lordo'):
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell.alignment = Alignment(vertical='center')

col_widths = {
    'Data emissione': 14, 'DenominazioneFornitore': 42, 'NumeroDocumento': 22,
    'TipoDocumento': 26,  'Imponibile': 14, 'Iva': 12, 'Lordo': 16, 'Categoria': 26,
}
for col_idx, header in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 15)

ws.freeze_panes = 'A2'
wb.save(OUTPUT_FILE)

print(f"\nFATTURE.xlsx generato: {len(result)} righe, {(result['Imponibile'] < 0).sum()} note di credito")
